"""
Google Maps Real Estate Lead Scraper (India)

Author  : Shankarsan Sahoo
Tech    : Python, Selenium, BeautifulSoup
Sources : Google Maps, Sulekha, TradeIndia

Features:
- Dynamic Google Maps scraping (no stale elements)
- Multi-source aggregation
- Phone extraction + validation
- Website filtering (removes directories like JustDial, 99acres)
- Excel export with formatting

Usage:
    python google_maps_real_estate_scraper_india.py

Output:
    hyderabad_realestate_leads.xlsx
"""

import time, re, logging, sys, warnings
warnings.filterwarnings("ignore")

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, TimeoutException

# ── Config ────────────────────────────────────────────────────────────────────
OUTPUT_FILE     = "hyderabad_realestate_leads.xlsx"
REQUEST_TIMEOUT = 6   # reduced — dead sites timeout fast now
SCROLL_PAUSE    = 2.0
COLS            = ["Company Name", "Phone", "Website", "Source", "Phone Found"]

GMAPS_QUERIES = [
    "real estate builders Hyderabad",
    "property developers Hyderabad",
    "real estate companies Hyderabad",
    "residential builders Hyderabad",
    "construction companies Hyderabad",
]

SULEKHA_URLS = [
    "https://www.sulekha.com/real-estate-agents/hyderabad",
    "https://www.sulekha.com/builders-and-developers/hyderabad",
]

TRADEINDIA_URLS = [
    "https://www.tradeindia.com/search.html?search_string=real+estate+builders+hyderabad",
    "https://www.tradeindia.com/search.html?search_string=property+developers+hyderabad",
]

BLACKLIST = {
    "google","youtube","facebook","twitter","instagram","linkedin",
    "reddit","quora","wikipedia","justdial","indiamart","99acres",
    "magicbricks","nobroker","housing","makaan","commonfloor",
    "scribd","slideshare","pinterest","sulekha","tradeindia",
    "exportersindia","grihashakti","blogger","blogspot","wordpress",
    "medium","mouthshut","amazon","flipkart","snapdeal","olx",
    "squarespace","wix","weebly","shopify","maps.google",
}

PHONE_PATTERNS = [
    r'\+91[\s\-]?\d{5}[\s\-]?\d{5}',        # +91 XXXXX XXXXX
    r'\+91[\s\-]?\d{10}',                        # +91XXXXXXXXXX
    r'\b[6-9]\d{4}[\s\-]?\d{5}\b',            # mobile with optional space
    r'\b[6-9]\d{9}\b',                            # mobile no space
    r'\b0\d{2}[\s\-]?\d{4}[\s\-]?\d{4}\b', # 040 2335 2708
    r'\b0\d{4,5}[\s\-]?\d{4,5}\b',            # 080999 72972, 090003 51269
    r'\b0\d{9,10}\b',                             # 09000351269 no spaces
]

CONTACT_PATHS = [
    "/contact-us", "/contact", "/contactus",
    "/about-us",   "/about",   "/reach-us",
    "/get-in-touch", "/enquiry",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────
def get_session():
    s = requests.Session()
    retry = Retry(total=1, backoff_factor=0.5, status_forcelist=[429,500,502,503,504])
    s.mount("http://",  HTTPAdapter(max_retries=retry))
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-IN,en;q=0.9",
    })
    return s


def build_driver(headless=False):
    opts = webdriver.ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--lang=en-IN")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    try:
        d = webdriver.Chrome(options=opts)
        d.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        return d
    except WebDriverException as e:
        log.error(f"ChromeDriver failed: {e}")
        sys.exit(1)


def dismiss_popups(driver):
    for sel in [
        "button[id*='accept']", "button[id*='agree']",
        "button[aria-label*='Accept']", "button[aria-label*='Reject']",
        "form[action*='consent'] button",
    ]:
        try:
            driver.find_element(By.CSS_SELECTOR, sel).click()
            time.sleep(0.4)
        except Exception:
            pass


def safe_get(session, url):
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 200:
            return r
    except requests.exceptions.SSLError:
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT, verify=False)
            if r.status_code == 200:
                return r
        except Exception:
            pass
    except Exception:
        pass
    return None


def extract_phones(text):
    phones = []
    for pat in PHONE_PATTERNS:
        for m in re.findall(pat, text):
            digits = re.sub(r"\D", "", m)
            if 10 <= len(digits) <= 12:
                phones.append(re.sub(r"\s", "", m).strip())
    return list(dict.fromkeys(phones))


def valid_phone(phone):
    if not phone:
        return False
    digits = re.sub(r"\D", "", phone)
    if not (8 <= len(digits) <= 12):
        return False
    if len(set(digits)) <= 2:
        return False
    if digits in ("1234567890", "9876543210", "0123456789"):
        return False
    # Indian mobile: 10 digits starting 6-9
    if len(digits) == 10 and digits[0] in "6789":
        return True
    # +91 prefix mobile: 91 + 10-digit mobile = 12 digits
    if len(digits) == 12 and digits[:2] == "91" and digits[2] in "6789":
        return True
    # Landline 11 digits starting 0 (e.g. 080999 72972, 090003 51269)
    if len(digits) == 11 and digits[0] == "0":
        return True
    # Landline 10 digits with known Indian STD prefix (04x, 08x, 07x, 06x)
    if len(digits) == 10 and digits[:2] in ("04", "08", "07", "06"):
        return True
    return False


def is_real_company_site(url):
    if not url or not url.startswith("http"):
        return False
    if "google.com" in url:
        return False
    try:
        domain = url.split("//")[-1].split("/")[0].lower()
        domain = re.sub(r"^www\d*\.", "", domain)
        base   = domain.split(".")[0]
        return base not in BLACKLIST
    except Exception:
        return False


def root_url(href):
    try:
        parts = href.split("/")
        return parts[0] + "//" + parts[2]
    except Exception:
        return href


def get_phone_from_site(session, base_url):
    urls_to_try = [base_url] + [base_url.rstrip("/") + p for p in CONTACT_PATHS]
    for url in urls_to_try:
        r = safe_get(session, url)
        if not r:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='tel:']"):
            raw   = a.get("href", "").replace("tel:", "").strip()
            phone = re.sub(r"[\s\-\(\)]", "", raw)
            if valid_phone(phone):
                return phone
        for p in extract_phones(r.text):
            if valid_phone(p):
                return p
    return ""


def make_record(name, phone, website, source):
    return {
        "Company Name": name.strip(),
        "Phone":        phone.strip(),
        "Website":      website.strip(),
        "Source":       source,
        "Phone Found":  "Yes" if phone.strip() else "No",
    }


def dedup_add(results, new):
    existing = {r["Company Name"].lower() for r in results}
    for r in new:
        if r["Company Name"].lower() not in existing:
            existing.add(r["Company Name"].lower())
            results.append(r)
    return results


# ── Google Maps phone extraction from loaded place page ───────────────────────
def extract_phone_from_page(driver):
    # From debug: all info rows are in div.AeaXub div.Io6YTe
    # texts = [address, website_domain, phone, plus_code]
    # Just scan ALL of them and return the first valid phone
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, "div.AeaXub div.Io6YTe"):
            txt = el.text.strip()
            if not txt:
                continue
            phones = extract_phones(txt)
            p = next((pp for pp in phones if valid_phone(pp)), "")
            if p:
                return p
    except Exception:
        pass
    return ""


def extract_website_from_page(driver):
    # Method 1: Website anchor href — DevTools confirmed: a.lcr4fd[data-value='Website']
    for sel in [
        "a.lcr4fd[data-value='Website']",
        "a[data-value='Website']",
        "a[aria-label*='website' i]",
    ]:
        try:
            el   = driver.find_element(By.CSS_SELECTOR, sel)
            href = el.get_attribute("href") or ""
            if "google.com/url" in href:
                m = re.search(r"[?&]q=([^&]+)", href)
                if m:
                    href = requests.utils.unquote(m.group(1))
            if href.startswith("http") and is_real_company_site(href):
                return root_url(href)
        except Exception:
            continue

    # Method 2: domain text from AeaXub > rogA2c.ITvuef > Io6YTe
    try:
        for el in driver.find_elements(
            By.CSS_SELECTOR,
            "div.AeaXub div.rogA2c.ITvuef div.Io6YTe, div.rogA2c.ITvuef div.Io6YTe"
        ):
            txt = el.text.strip()
            if txt and "." in txt and " " not in txt:
                candidate = "https://" + txt
                if is_real_company_site(candidate):
                    return candidate
    except Exception:
        pass

    return ""


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 1 — Google Maps
# Key fix: collect all place page URLs first, then navigate to each one directly.
# Each driver.get() loads a completely fresh DOM — zero stale element problems.
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_google_maps(driver, session):
    log.info("═══ [1/3] Google Maps ═══")
    results    = []
    place_urls = []  # collect all URLs before visiting any

    # ── Step 1: collect place URLs from search results ────────────────────────
    for query in GMAPS_QUERIES:
        log.info(f"  Collecting URLs for: {query}")
        try:
            url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/@17.385,78.4867,13z"
            driver.get(url)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']"))
            )
            dismiss_popups(driver)
            time.sleep(2)

            # Scroll to load all results
            feed = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
            last_count = 0
            for _ in range(8):
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", feed)
                time.sleep(SCROLL_PAUSE)
                cards = driver.find_elements(By.CSS_SELECTOR, "div[role='feed'] > div > div > a")
                if len(cards) == last_count:
                    break
                last_count = len(cards)

            # Grab all place URLs from cards
            cards = driver.find_elements(By.CSS_SELECTOR, "div[role='feed'] > div > div > a")
            for card in cards:
                try:
                    href = card.get_attribute("href") or ""
                    if href and href not in place_urls:
                        place_urls.append(href)
                except Exception:
                    pass
            log.info(f"  {len(cards)} listings found for '{query}'")

        except Exception as e:
            log.warning(f"  Failed to collect URLs for '{query}': {e}")

    log.info(f"  Total unique place URLs collected: {len(place_urls)}")

    # ── Step 2: visit each place URL directly — fresh DOM every time ──────────
    for idx, place_url in enumerate(place_urls):
        try:
            driver.get(place_url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1.DUwDvf"))
            )
            time.sleep(1.2)

            # Name
            name = ""
            for sel in ["h1.DUwDvf", "h1[class*='fontHeadlineLarge']"]:
                try:
                    name = driver.find_element(By.CSS_SELECTOR, sel).text.strip()
                    if name:
                        break
                except Exception:
                    pass
            if not name:
                continue

            # Website and phone — fresh DOM, no stale elements possible
            website = extract_website_from_page(driver)

            phone   = extract_phone_from_page(driver)

            if name and website:
                log.info(f"  [{idx+1}/{len(place_urls)}] ✓ {name} | {phone or '—'} | {website}")
                results.append(make_record(name, phone, website, "Google Maps"))
            else:
                log.debug(f"  [{idx+1}] ✗ '{name}' — no website")

            time.sleep(0.5)

        except TimeoutException:
            log.debug(f"  [{idx+1}] Timeout")
            continue
        except Exception as e:
            log.debug(f"  [{idx+1}] Error: {e}")
            continue

    log.info(f"  Google Maps → {len(results)} records")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 2 — Sulekha
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_sulekha(driver, session):
    log.info("═══ [2/3] Sulekha ═══")
    results = []

    for url in SULEKHA_URLS:
        log.info(f"  URL: {url}")
        try:
            driver.get(url)
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            dismiss_popups(driver)
            time.sleep(2)

            for _ in range(5):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(SCROLL_PAUSE)

            soup  = BeautifulSoup(driver.page_source, "html.parser")
            cards = (
                soup.select("div.card-body") or
                soup.select("div.listing-card") or
                soup.select("[class*='biz-listing']") or
                soup.select("li.listing") or
                soup.select("div.srp-card") or []
            )
            log.info(f"  Found {len(cards)} cards")

            for card in cards:
                try:
                    name_el = (
                        card.select_one("h2") or
                        card.select_one("h3") or
                        card.select_one("a.biz-name") or
                        card.select_one("[class*='name']")
                    )
                    name = name_el.get_text(strip=True) if name_el else ""
                    if not name:
                        continue

                    phone   = ""
                    website = ""

                    phones = extract_phones(card.get_text())
                    phone  = next((p for p in phones if valid_phone(p)), "")

                    for a in card.select("a[href]"):
                        href = a.get("href", "")
                        if href.startswith("http") and is_real_company_site(href):
                            website = root_url(href)
                            break

                    if not website:
                        detail_link = card.select_one("a[href*='/hyderabad/']")
                        if detail_link:
                            detail_url = detail_link.get("href", "")
                            if not detail_url.startswith("http"):
                                detail_url = "https://www.sulekha.com" + detail_url
                            r = safe_get(session, detail_url)
                            if r:
                                dsoup = BeautifulSoup(r.text, "html.parser")
                                for a in dsoup.select("a[href]"):
                                    href = a.get("href", "")
                                    if href.startswith("http") and is_real_company_site(href):
                                        website = root_url(href)
                                        break
                                if not phone:
                                    ph = extract_phones(r.text)
                                    phone = next((p for p in ph if valid_phone(p)), "")


                    if name and website:
                        log.info(f"    ✓ {name} | {phone or '—'} | {website}")
                        results.append(make_record(name, phone, website, "Sulekha"))

                except Exception as e:
                    log.debug(f"    Card error: {e}")

        except Exception as e:
            log.warning(f"  Sulekha failed {url}: {e}")

    log.info(f"  Sulekha → {len(results)} records")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE 3 — TradeIndia
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_tradeindia(session):
    log.info("═══ [3/3] TradeIndia ═══")
    results = []

    for url in TRADEINDIA_URLS:
        log.info(f"  URL: {url}")
        try:
            r = safe_get(session, url)
            if not r:
                continue
            soup  = BeautifulSoup(r.text, "html.parser")
            cards = (
                soup.select("div.companyDetails") or
                soup.select("div.company-detail") or
                soup.select("li.listing-item") or
                soup.select("[class*='company']") or []
            )
            log.info(f"  Found {len(cards)} cards")

            for card in cards:
                try:
                    name_el = (
                        card.select_one("h3.companyName") or
                        card.select_one("a.company-name") or
                        card.select_one("h2") or
                        card.select_one("h3")
                    )
                    name = name_el.get_text(strip=True) if name_el else ""
                    if not name:
                        continue

                    phones = extract_phones(card.get_text())
                    phone  = next((p for p in phones if valid_phone(p)), "")

                    website = ""
                    for a in card.select("a[href]"):
                        href = a.get("href", "")
                        if href.startswith("http") and is_real_company_site(href):
                            website = root_url(href)
                            break

                    if not website:
                        profile_a = card.select_one("a[href*='tradeindia']")
                        if profile_a:
                            profile_url = profile_a.get("href", "")
                            if not profile_url.startswith("http"):
                                profile_url = "https://www.tradeindia.com" + profile_url
                            pr = safe_get(session, profile_url)
                            if pr:
                                psoup = BeautifulSoup(pr.text, "html.parser")
                                for a in psoup.select("a[href]"):
                                    href = a.get("href", "")
                                    if href.startswith("http") and is_real_company_site(href):
                                        website = root_url(href)
                                        break
                                if not phone:
                                    ph = extract_phones(pr.text)
                                    phone = next((p for p in ph if valid_phone(p)), "")


                    if name and website:
                        log.info(f"    ✓ {name} | {phone or '—'} | {website}")
                        results.append(make_record(name, phone, website, "TradeIndia"))

                except Exception as e:
                    log.debug(f"    Card error: {e}")

            time.sleep(1)
        except Exception as e:
            log.warning(f"  TradeIndia failed {url}: {e}")

    log.info(f"  TradeIndia → {len(results)} records")
    return results


# ── Excel Output ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
YES_FILL    = PatternFill("solid", fgColor="C6EFCE")
NO_FILL     = PatternFill("solid", fgColor="FFEB9C")
THIN        = Side(style="thin")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COL_WIDTHS  = {"Company Name": 38, "Phone": 22, "Website": 48, "Source": 14, "Phone Found": 14}


def save_to_excel(records, filename):
    if not records:
        records = [make_record("No leads found", "", "", "N/A")]

    df = (
        pd.DataFrame(records, columns=COLS)
        .drop_duplicates(subset=["Company Name"])
        .reset_index(drop=True)
    )

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Leads", index=False)

    wb = load_workbook(filename)
    ws = wb["All Leads"]
    ws.delete_rows(1, ws.max_row)

    for ci, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci)
        c.value, c.font, c.fill = h, HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 16)

    for ri, row in df.iterrows():
        has_phone = str(row.get("Phone Found", "")).strip() == "Yes"
        row_fill  = YES_FILL if has_phone else NO_FILL
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri + 2, column=ci)
            c.value     = str(val) if val else ""
            c.font      = Font(
                name="Arial", size=10, bold=(ci == 5),
                color=("375623" if has_phone else "9C5700") if ci == 5 else "000000"
            )
            c.fill      = row_fill
            c.border    = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

    with_phone    = len(df[df["Phone Found"] == "Yes"])
    without_phone = len(df[df["Phone Found"] == "No"])
    log.info(f"✅  {len(df)} leads saved → {filename}")
    log.info(f"    🟢 {with_phone} with phone   🟡 {without_phone} website only")
    return len(df)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log.info("🚀  Hyderabad Real Estate Lead Scraper")
    log.info("    Sources : Google Maps + Sulekha + TradeIndia")
    log.info("    Fix     : Navigate directly to each place URL — no stale DOM\n")

    session = get_session()
    driver  = build_driver(headless=False)
    results = []

    try:
        gmaps    = scrape_google_maps(driver, session)
        results  = dedup_add(results, gmaps)

        sulekha  = scrape_sulekha(driver, session)
        results  = dedup_add(results, sulekha)

        driver.quit()

        tradeindia = scrape_tradeindia(session)
        results    = dedup_add(results, tradeindia)

    except Exception as e:
        log.error(f"Fatal: {e}")
        try:
            driver.quit()
        except Exception:
            pass

    total      = save_to_excel(results, OUTPUT_FILE)
    with_phone = sum(1 for r in results if r["Phone"])

    print("\n" + "═" * 58)
    print(f"  ✅  {total} unique leads saved → {OUTPUT_FILE}")
    print(f"  🟢  {with_phone} with phone")
    print(f"  🟡  {total - with_phone} website only")
    print(f"  📋  Company Name | Phone | Website | Source | Phone Found")
    print("═" * 58)


if __name__ == "__main__":
    main()